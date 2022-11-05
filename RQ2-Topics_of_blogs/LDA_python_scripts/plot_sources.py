import matplotlib.pyplot as plt
import openpyxl
import numpy as np
import pandas as pd

TOPIC_COUNT = 7

# loads data for the topic relevance box plot
def load_box_data(filename):
    array2d = []
    for _ in range(TOPIC_COUNT):
        array2d.append([])
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows():
        if (isinstance(row[4].value, (int, float))):
            array2d[row[0].value - 1].append(row[4].value)    
    return np.array(array2d)

# merges several subcategory sourcetypes to the one source
def match_general_type(typename):
    if typename.startswith("Community blog"):
        return "Community blog" # matches subtypes
    if typename == "Technology specific community blog":
        return "Community blog"
    if typename.startswith("Technology Vendor") or typename.startswith("Technology vendor"):
        return "Technology Vendor"
    else:
        return typename

# loads the data necessary to create a discete bubble chart to match the origin to topic
def load_bubble_data(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    names = {'Technology Vendor: eCommerce', 'Technology Vendor: authentication', 'Technology vendor', 
        'Technology Vendor: chatbots', 'Technology Vendor: healthcare', 'Technology vendor: finance', 
        'Technology Vendor: integration', 'Tutorial', 'University blog', 'Personal', 'IT Service Company', 'Community blog on a tutorial site', 
        'Educational IT course provider', 'Technology Vendor: Finance', 'Magazines and Newspapers', 'Technology Vendor: analytics', 'Technology specific community blog', 
        'Community blog on Educational IT course provider site', 'Community blog', 'Technology Vendor: books', 
        'Technology Vendor: API', 'Technology Vendor', 'Technology Vendor: cloud'}
    typenames_set = set()
    for name in names:
        typenames_set.add(match_general_type(name))
    typenames = list(typenames_set)
    dimensions = len(typenames) * TOPIC_COUNT

    name_topic_dict = dict()

    #fill topic source counts
    for x in range(dimensions):
        name_topic_dict[typenames[x % len(typenames)] + str( x % TOPIC_COUNT + 1)] = 0
    
    for row in ws.iter_rows():
        if (row[5].value and row[5].value in names):
            name_topic_dict[match_general_type(row[5].value)+str(row[0].value)] += 1
    for x, y in name_topic_dict.items():
        print(x, y)
    
    SE = dict()
    DL = dict()
    Count = dict()

    #lets create a panda dataframe
    for x, item in enumerate(name_topic_dict.values()):
        SE[x] = typenames[x % len(typenames)]
        DL[x] = x % 7 + 1
        Count[x] = item
    df = pd.DataFrame({'SE' : SE, 'DL' : DL, 'Count' : Count})
    return df

def plot_relevance(data):    
    fig, ax = plt.subplots()
    #pos = np.arange(len(data)) + 1
    bp = ax.boxplot(data)

    ax.set_xlabel('Topic number')
    ax.set_ylabel('Relevance')    
    plt.show()

def plot_topic_types(df):
    s_scaling = 20 # modifies the bubble scale
    fig = plt.figure(figsize= (22,12))
    ax = fig.add_subplot(111)
    ax.scatter(x="DL", y="SE", s=df.Count*s_scaling,  # scaling the size here
        data=df, alpha=0.7, c=df.Count*5000)
    ax.set_xlabel('Topic number', fontweight='bold')
    ax.set_ylabel('Blog source', fontweight='bold')
    plt.gcf().subplots_adjust(left=0.3)
    plt.show()
    

def main():
    pdata = load_bubble_data("match_topics_types.xlsx")
    plot_topic_types(pdata)
    data = load_box_data("match_topics_types.xlsx")
    plot_relevance(data)    
    return

if __name__ == "__main__":
    main()