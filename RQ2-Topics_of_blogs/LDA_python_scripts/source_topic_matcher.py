import openpyxl

topic_url = "topics_to_url.xlsx"
source_url = "BlogTypes.xlsx"

def main():
    wb_topics = openpyxl.load_workbook(topic_url)
    ws_t = wb_topics.active
    
    wb_sources = openpyxl.load_workbook(source_url)
    source_map = {}
    total_count = 0
    ws_s = wb_sources.active
    for row in  ws_s.iter_rows(min_row=1, max_col=3, values_only=True):
        source_map[row[0]] = [row[1], row[2]]       
    for index, row in  enumerate(ws_t.iter_rows(min_row=1, max_col=6, values_only=True), 1):
        if row[1] in source_map.keys():
            match = source_map[row[1]]
            print(match[0], match[1])
            total_count += 1
            ws_t.cell(row = index, column = 5).value = match[0]
            ws_t.cell(row = index, column = 6).value = match[1]            
    print(total_count)
    wb_topics.save("match_topics_types.xlsx")
    return


if __name__ == "__main__":
    main()