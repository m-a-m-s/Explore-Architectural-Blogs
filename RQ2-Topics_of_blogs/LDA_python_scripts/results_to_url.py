import os
import sys
import openpyxl


def match_file_to_url(f, uf):
    for [url, filename] in uf:
        if filename == f:
            print(url + " = " + filename)
            return url
    return "None: " + f


def main():
    # destination file name
    results_file_name = "topics_to_url.xlsx"
    if len(sys.argv) == 2 and sys.argv[1].endswith(".xlsx"):
        results_file_name = sys.argv[1]

    # tuple of [url, filename]
    url_filename = []
    match_filename = []

    # import.py results containing: url, returncode, saved filename
    wb_pages = openpyxl.load_workbook("import_results.xlsx")
    ws = wb_pages.active
    # LDA.py results containing: topic number, file name, score
    wb_results = openpyxl.load_workbook("lda_result_names.xlsx")    
    ws2 = wb_results.active

    txtfiles_path = os.path.join('txtfiles')

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        url_filename.append([row[0], row[2].replace(txtfiles_path, '')])
    
    txtfiles__cat1_path = os.path.join('txtfiles', 'category_1_folder')

    for row in ws2.iter_rows(min_row=1, max_col=3, values_only=True):
        match_filename.append([row[0], row[1].replace(txtfiles__cat1_path, ''), row[2]])

    print(len(url_filename))
    print(url_filename[10])
    print(len(match_filename))
    print(match_filename[10])

    wb_final = openpyxl.Workbook()
    ws3 = wb_final.active
    for index, [match, filename, score] in enumerate(sorted(match_filename)):
        print(filename)
        ws3.cell(row = index +1, column = 1).value = match
        ws3.cell(row = index +1, column = 2).value = match_file_to_url(filename, url_filename)
        ws3.cell(row = index +1, column = 3).value = filename
        ws3.cell(row = index +1, column = 4).value = score 
    wb_final.save(results_file_name)

    url_filename.sort(key= lambda x: x[1])

    #for [u, f] in url_filename:
        #print(f)

if __name__ == "__main__":
    main()