import sys
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import os
from tqdm import tqdm

HTML_FOLDER = "htmlfiles"
HEADERS = ""
REDIR = False
DOCUMENTS_FOLDER = "txtfiles"


def url_to_name(s):
    # Remove all non-word characters (everything except numbers and letters)
    s = re.sub(r"[^\w\s]", '', s)

    # Replace all runs of whitespace with a single dash
    s = re.sub(r"\s+", '-', s)
    return s

def get_name(url):
    name = url.rsplit('/', 1)[-1]
    file_name = url_to_name(name)
    file_location = os.path.join(DOCUMENTS_FOLDER, file_name+"_full.txt")
    return file_location

# reads urls from excel sheet
def read_urls(filename): 
    wb = openpyxl.load_workbook(filename)
    print(wb.sheetnames)
    ws = wb.active  
    for index, row in enumerate (ws.iter_rows(min_row = 2)):
        url = row[0].value      
        result = copy_page(url)
        if result[1] == 200:
            name = write_page(result[0], url)
        else:
            name = get_name(url)
        ws.cell(row = index + 2, column = 2).value = result[1]
        ws.cell(row = index + 2, column = 3).value = name
    wb.save('import_results.xlsx')

# downloads copy of webpage
def copy_page(url):
    print(url)
    if not HEADERS:
        headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
    else:
        headers = {'User-Agent': HEADERS}
    try:
        result = requests.get(url, headers=headers, allow_redirects=REDIR)       
        if result.status_code == 200:
            return [result, result.status_code]
        else:
            return [None, result.status_code]
    except:
        print("failed to load " + url)        
        return [None, -1]

# removes html elements, retaining only the text from webpage    
def write_page(page, url):
    if page is None:
        return

    soup = BeautifulSoup(page.content, "html.parser")

    # finds the title of blog to be used as filename
    # alternative if not available generate one using urlify func
    title = soup.find('title')    
    name = ""
    if title:
        name = title.text
    else:
        name = url.rsplit('/', 1)[-1]
    file_name = url_to_name(name)    

    # Create a local copy of the html file
    html_path = os.path.join(HTML_FOLDER, file_name+"_full.html")
    with open(html_path, "w", encoding="utf-8") as f:
        print(html_path)
        f.write(page.text)      
        f.close()
    
    # Save the txt contents of the html file
    txt_path = os.path.join(DOCUMENTS_FOLDER, file_name+"_full.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(soup.get_text())      
        f.close()

    return txt_path

# clean up existing html files to text files
def cleanup_existing(file_name):

    html_path = os.path.join(HTML_FOLDER, file_name)
    file = open(html_path, 'r', encoding='utf-8')
    data = file.read()
    soup = BeautifulSoup(data, "html.parser")
    #new_name = file_name.replace(".html", ".txt") # Original solution, doesn't work with .htm 
    new_name = os.path.splitext(file_name)[0] + '.txt'    
    txt_path = os.path.join(DOCUMENTS_FOLDER, new_name)
    with open(txt_path, "w+", encoding="utf-8") as f:
        f.write(soup.get_text())      
        f.close()
    file.close()

def usage():
    print("all args are optional, without input file in arg[1], clean up on htmlfiles performed")
    print("arg[1] = .xlxs file with list of urls starting from row 2")
    print("arg[2] is user-argent header for requests library GET")
    print("arg[3] \"yes\" for allow redir")

if __name__ == "__main__":
    args = sys.argv
    # use GET to load all listed blogs
    if len(args) == 4:
        REDIR = args[3] == "yes"
    if len(args) >= 3:
        HEADERS = args[2]
    if len(args) >= 2:
        read_urls(args[1])

    # parse existing local copies
    else:
        for file_name in tqdm(os.listdir("htmlfiles")):
            if file_name.endswith(".html") or file_name.endswith(".htm"):
                cleanup_existing(file_name)